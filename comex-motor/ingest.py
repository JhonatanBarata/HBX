# -*- coding: utf-8 -*-
"""
HBX Comex — N1: ingestão do bulk Comex Stat + cadastro histórico de empresas
para o banco analítico local (comex.duckdb). Isolado do Postgres do CRM.

Fontes:
  data/raw/tabelas/*.csv   — auxiliares oficiais (latin-1, ';')
  data/raw/mun/*.csv       — SH4 x município x país, mensal
  data/raw/ncm/*.csv       — NCM x UF x via x URF, mensal
  data/raw/cadastro/*.xlsx — lista oficial de empresas EXP/IMP (2018-2020,
                             resgatada do Wayback; publicação foi descontinuada
                             pela SECEX em mar/2023)
"""
import glob
import os
import re
import sys

import duckdb
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
RAW = os.path.join(HERE, "data", "raw")
DB = os.path.join(HERE, "data", "comex.duckdb")


def load_aux(con):
    tabelas = {
        "aux_ncm": "NCM.csv",
        "aux_ncm_sh": "NCM_SH.csv",
        "aux_pais": "PAIS.csv",
        "aux_pais_bloco": "PAIS_BLOCO.csv",
        "aux_uf": "UF.csv",
        "aux_uf_mun": "UF_MUN.csv",
        "aux_via": "VIA.csv",
        "aux_urf": "URF.csv",
    }
    for table, fname in tabelas.items():
        path = os.path.join(RAW, "tabelas", fname).replace("\\", "/")
        con.execute(f"""
            CREATE OR REPLACE TABLE {table} AS
            SELECT * FROM read_csv('{path}', delim=';', header=true,
                                   encoding='latin-1', quote='"', all_varchar=true,
                                   strict_mode=false)
        """)
        n = con.execute(f"SELECT count(*) FROM {table}").fetchone()[0]
        print(f"{table}: {n} linhas")


def load_flows(con):
    # IMP tem VL_FRETE/VL_SEGURO (componentes do CIF); EXP não. Normalizamos com NULL.
    base_ncm = ("CO_ANO INT, CO_MES INT, CO_NCM VARCHAR, CO_UNID VARCHAR, "
                "CO_PAIS VARCHAR, SG_UF_NCM VARCHAR, CO_VIA VARCHAR, CO_URF VARCHAR, "
                "QT_ESTAT BIGINT, KG_LIQUIDO BIGINT, VL_FOB BIGINT")
    specs = {
        "mun": {
            "IMP": ("CO_ANO INT, CO_MES INT, SH4 VARCHAR, CO_PAIS VARCHAR, "
                    "SG_UF_MUN VARCHAR, CO_MUN VARCHAR, KG_LIQUIDO BIGINT, VL_FOB BIGINT", ""),
            "EXP": ("CO_ANO INT, CO_MES INT, SH4 VARCHAR, CO_PAIS VARCHAR, "
                    "SG_UF_MUN VARCHAR, CO_MUN VARCHAR, KG_LIQUIDO BIGINT, VL_FOB BIGINT", ""),
        },
        "ncm": {
            "IMP": (base_ncm + ", VL_FRETE BIGINT, VL_SEGURO BIGINT", ""),
            "EXP": (base_ncm, ", NULL::BIGINT AS VL_FRETE, NULL::BIGINT AS VL_SEGURO"),
        },
    }
    for kind, by_fluxo in specs.items():
        frames = []
        for path in sorted(glob.glob(os.path.join(RAW, kind, "*.csv"))):
            fluxo = "IMP" if os.path.basename(path).startswith("IMP") else "EXP"
            cols, extra = by_fluxo[fluxo]
            p = path.replace("\\", "/")
            frames.append(
                f"SELECT '{fluxo}' AS fluxo, *{extra} FROM read_csv('{p}', delim=';', "
                f"header=true, encoding='latin-1', columns={{{_cols_dict(cols)}}})"
            )
        con.execute(f"CREATE OR REPLACE TABLE flow_{kind} AS " + " UNION ALL ".join(frames))
        n = con.execute(f"SELECT count(*) FROM flow_{kind}").fetchone()[0]
        print(f"flow_{kind}: {n} linhas")


def _cols_dict(cols: str) -> str:
    pairs = []
    for part in cols.split(","):
        name, typ = part.strip().split(" ", 1)
        pairs.append(f"'{name}': '{typ}'")
    return ", ".join(pairs)


def load_cadastro(con):
    con.execute("""
        CREATE OR REPLACE TABLE cadastro_empresas (
            ano INT, fluxo VARCHAR, cnpj VARCHAR, empresa VARCHAR,
            endereco VARCHAR, numero VARCHAR, bairro VARCHAR, cep VARCHAR,
            municipio VARCHAR, uf VARCHAR, cnae VARCHAR, natureza VARCHAR
        )
    """)
    for path in sorted(glob.glob(os.path.join(RAW, "cadastro", "*.xlsx"))):
        ano = int(re.search(r"(\d{4})", os.path.basename(path)).group(1))
        wb = openpyxl.load_workbook(path, read_only=True)
        for sheet in wb.sheetnames:
            fluxo = "EXP" if sheet.startswith("EXP") else "IMP"
            ws = wb[sheet]
            ws.reset_dimensions()
            rows, started = [], False
            for row in ws.iter_rows(values_only=True):
                if not started:
                    started = bool(row) and str(row[0]).strip() == "CNPJ"
                    continue
                if not row or row[0] is None:
                    continue
                vals = [str(c).strip() if c is not None else None for c in row[:10]]
                vals += [None] * (10 - len(vals))
                cnpj = re.sub(r"\D", "", vals[0] or "")
                if len(cnpj) != 14:
                    continue
                rows.append([ano, fluxo, cnpj] + vals[1:])
            con.executemany(
                "INSERT INTO cadastro_empresas VALUES (?,?,?,?,?,?,?,?,?,?,?,?)", rows
            )
            print(f"cadastro {ano} {fluxo}: {len(rows)} linhas")
    # visão 1-linha-por-empresa/fluxo com o último ano em que apareceu na lista oficial
    con.execute("""
        CREATE OR REPLACE VIEW cadastro_atual AS
        SELECT fluxo, cnpj,
               arg_max(empresa, ano)   AS empresa,
               arg_max(municipio, ano) AS municipio,
               arg_max(uf, ano)        AS uf,
               arg_max(cnae, ano)      AS cnae,
               arg_max(natureza, ano)  AS natureza,
               max(ano)                AS ultimo_ano_na_lista,
               count(DISTINCT ano)     AS anos_na_lista
        FROM cadastro_empresas GROUP BY fluxo, cnpj
    """)


def main():
    os.makedirs(os.path.dirname(DB), exist_ok=True)
    con = duckdb.connect(DB)
    load_aux(con)
    load_flows(con)
    load_cadastro(con)
    con.close()
    print("OK ->", DB)


if __name__ == "__main__":
    sys.exit(main())
